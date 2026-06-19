"""데이터 에디터 탭 — session.bin (60B×N)을 테이블로 파싱·표시하고
더블클릭으로 필드를 편집한 뒤 재패킹(HMAC 재서명)해 저장한다."""
import struct
import hmac as _hmac
import hashlib
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import _src_path; _src_path.add()
from core_module import config
from core_module.v2v import PACKET_LEN, fmt_ms_of_day
from messages import Role, DriveBehavior

_FMT = '!BBBHdIBBffx'          # 28B body (v2v.py 와 동일)
_HDR = struct.calcsize(_FMT)    # 28

_ROLE_NAMES   = {r.value: r.name for r in Role}
_BEH_NAMES    = {b.value: b.name for b in DriveBehavior}
_ROLE_BY_NAME = {r.name: r for r in Role}
_BEH_BY_NAME  = {b.name: b for b in DriveBehavior}

# (id, header, px_width, anchor)
_COLS = [
    ('idx',      '#',             42,  'e'),
    ('tx_time',  '시각(tx_abs)', 145, 'center'),   # 파싱값 + \n + 4B hex
    ('role',     '역할',         105, 'center'),   # 파싱값 + \n + 1B hex
    ('seq',      'seq',           85, 'e'),         # 파싱값 + \n + 2B hex
    ('lane',     '차선',          65, 'center'),   # 파싱값 + \n + 1B hex
    ('behavior', '행동',         110, 'center'),   # 파싱값 + \n + 1B hex
    ('throttle', 'throttle_pwm', 125, 'e'),        # 파싱값 + \n + 4B hex
    ('steer',    'steer_pwm',    125, 'e'),        # 파싱값 + \n + 4B hex
    ('raw60', 'ver│typ│role│seq(2)│t_tx(8)│tx_abs(4)│lane│beh│thr(4)│str(4)│pad│HMAC(32)',
     1100, 'w'),
]

# 28B body 내 각 필드의 바이트 슬라이스
_SLICES = {
    'ver':          (0,  1),
    'type':         (1,  2),
    'role':         (2,  3),
    'seq':          (3,  5),
    't_tx':         (5,  13),
    'tx_abs':       (13, 17),
    'lane':         (17, 18),
    'behavior':     (18, 19),
    'throttle_pwm': (19, 23),
    'steer_pwm':    (23, 27),
    'padding':      (27, 28),
}


def _h(raw, a, b):
    """raw bytes[a:b] → 소문자 hex 문자열 (스페이스 구분)."""
    return raw[a:b].hex(' ')


def _fmt_raw60(raw):
    """60B → 필드 경계(│)로 구분된 hex 문자열. 로직 아날라이저 스타일."""
    groups = [
        raw[0:1],    # ver
        raw[1:2],    # type
        raw[2:3],    # role
        raw[3:5],    # seq  (2B)
        raw[5:13],   # t_tx (8B)
        raw[13:17],  # tx_abs (4B)
        raw[17:18],  # lane
        raw[18:19],  # behavior
        raw[19:23],  # throttle (4B)
        raw[23:27],  # steer    (4B)
        raw[27:28],  # padding
        raw[28:60],  # HMAC-SHA256 (32B)
    ]
    return ' │ '.join(g.hex(' ') for g in groups)


# ── 코덱 헬퍼 ────────────────────────────────────────────────────────────

def _parse_one(raw_bytes):
    """60B raw → (fields_dict | None, error_str | None).  key 검증 포함."""
    if len(raw_bytes) != PACKET_LEN:
        return None, f'길이 오류 {len(raw_bytes)}≠{PACKET_LEN}'
    body, mac = raw_bytes[:_HDR], raw_bytes[_HDR:]
    key = config.load_key()
    expected  = _hmac.new(key, body, hashlib.sha256).digest()
    hmac_ok   = _hmac.compare_digest(mac, expected)
    ver, typ, role, seq, t_tx, tx_abs, lane, beh, thr, st = struct.unpack(_FMT, body)
    return {
        'ver': ver, 'type': typ,
        'role': role, 'seq': seq, 't_tx': t_tx, 'tx_abs': tx_abs,
        'lane': lane, 'behavior': beh,
        'throttle_pwm': thr, 'steer_pwm': st,
        'hmac_ok': hmac_ok,
    }, None


def _repack(fields):
    """fields dict → 새 60B (body struct.pack + HMAC 재서명)."""
    body = struct.pack(
        _FMT,
        int(fields['ver']), int(fields['type']),
        int(fields['role']),
        int(fields['seq']) & 0xFFFF,
        float(fields['t_tx']),
        int(fields['tx_abs']),
        int(fields['lane']),
        int(fields['behavior']),
        float(fields['throttle_pwm']),
        float(fields['steer_pwm']),
    )
    key = config.load_key()
    return body + _hmac.new(key, body, hashlib.sha256).digest()


# ── 메인 탭 ──────────────────────────────────────────────────────────────

class DataEditorTab(ttk.Frame):
    """session.bin 파일을 로드해 패킷 테이블로 표시하고 더블클릭으로 편집한다."""

    def __init__(self, parent):
        super().__init__(parent)
        self._packets = []   # list of {'raw': bytes, 'fields': dict|None, 'error': str|None, 'edited': bool}
        self._path = None
        self._build_ui()

    # ── UI 구성 ──────────────────────────────────────────────────────────

    def _build_ui(self):
        # 상단 툴바
        bar = tk.Frame(self, bg='#f0f0f0', pady=3)
        bar.pack(fill='x', padx=6, pady=(4, 0))

        ttk.Button(bar, text='파일 열기', command=self._open_file).pack(side='left', padx=(0, 4))
        ttk.Separator(bar, orient='vertical').pack(side='left', fill='y', padx=4)
        tk.Label(bar, text='내보내기', bg='#f0f0f0', fg='#555').pack(side='left')
        ttk.Button(bar, text='CSV',  command=self._export_csv ).pack(side='left', padx=2)
        ttk.Button(bar, text='XLSX', command=self._export_xlsx).pack(side='left', padx=2)
        ttk.Button(bar, text='PDF',  command=self._export_pdf ).pack(side='left', padx=2)
        ttk.Button(bar, text='저장',          command=self._save_file).pack(side='right', padx=2)
        ttk.Button(bar, text='다른 이름으로 저장', command=self._save_as).pack(side='right', padx=2)

        self._path_var = tk.StringVar(value='session.bin 파일을 선택하세요')
        tk.Label(bar, textvariable=self._path_var, fg='#333', bg='#f0f0f0',
                 anchor='w').pack(side='left', fill='x', expand=True, padx=6)

        # 트리뷰 영역
        frm = tk.Frame(self)
        frm.pack(fill='both', expand=True, padx=6, pady=(4, 0))

        # 로직 아날라이저 스타일: 2줄 행 (파싱값 + hex bytes)
        _s = ttk.Style()
        _s.configure('DE.Treeview',
                     rowheight=46,
                     font=('Consolas', 10))
        _s.configure('DE.Treeview.Heading',
                     font=('TkDefaultFont', 9, 'bold'))

        col_ids = [c[0] for c in _COLS]
        self._tree = ttk.Treeview(frm, columns=col_ids, show='headings',
                                  selectmode='browse', style='DE.Treeview')
        for cid, label, width, anchor in _COLS:
            self._tree.heading(cid, text=label)
            self._tree.column(cid, width=width, minwidth=30, anchor=anchor,
                              stretch=(cid == 'raw60'))

        vsb = ttk.Scrollbar(frm, orient='vertical',   command=self._tree.yview)
        hsb = ttk.Scrollbar(frm, orient='horizontal', command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)

        # 색 태그
        self._tree.tag_configure('error',    background='#ffe0e0', foreground='#990000')
        self._tree.tag_configure('hmac_bad', background='#ffd0a0', foreground='#885500')
        self._tree.tag_configure('edited',   background='#fffacc', foreground='#444400')
        # 정상 행 교대색 (줄 구분)
        self._tree.tag_configure('evenrow',  background='#ffffff')
        self._tree.tag_configure('oddrow',   background='#eef1f5')

        self._tree.bind('<Double-1>', self._on_double_click)

        # 마우스 휠 스크롤 (Windows)
        self._tree.bind('<Enter>', lambda _: self._tree.bind_all(
            '<MouseWheel>', lambda ev: self._tree.yview_scroll(int(-1 * ev.delta / 120), 'units')))
        self._tree.bind('<Leave>', lambda _: self._tree.unbind_all('<MouseWheel>'))

        # 상태 바
        self._status_var = tk.StringVar(value='파일을 열면 패킷 목록이 표시됩니다')
        tk.Label(self, textvariable=self._status_var, anchor='w', fg='#555',
                 bg='#e8e8e8', relief='sunken', padx=6).pack(fill='x', side='bottom')

    # ── 파일 입출력 ──────────────────────────────────────────────────────

    def _open_file(self):
        path = filedialog.askopenfilename(
            filetypes=[('Binary', '*.bin'), ('All', '*.*')],
            title='session.bin 선택')
        if path:
            self._load(path)

    def _load(self, path):
        self._status_var.set('로딩 중...')
        self.update_idletasks()
        try:
            data = open(path, 'rb').read()
        except OSError as e:
            messagebox.showerror('오류', str(e))
            return

        self._path = path
        self._path_var.set(path)
        self._packets.clear()
        n_err = n_hmac = 0

        for i in range(0, len(data), PACKET_LEN):
            raw = data[i:i + PACKET_LEN]
            if len(raw) < PACKET_LEN:
                break
            fields, err = _parse_one(raw)
            self._packets.append({'raw': raw, 'fields': fields,
                                   'error': err, 'edited': False})
            if err:
                n_err += 1
            elif not fields['hmac_ok']:
                n_hmac += 1

        self._refresh_tree()
        parts = [f'{len(self._packets)}개 패킷 로드']
        if n_err:
            parts.append(f'파싱 오류 {n_err}개')
        if n_hmac:
            parts.append(f'HMAC 불일치 {n_hmac}개')
        parts.append('더블클릭으로 편집')
        self._status_var.set(' | '.join(parts))

    def _save_file(self):
        if not self._path:
            self._save_as()
            return
        self._write(self._path)

    def _save_as(self):
        if not self._packets:
            messagebox.showinfo('저장', '로드된 패킷이 없습니다.')
            return
        path = filedialog.asksaveasfilename(
            defaultextension='.bin',
            filetypes=[('Binary', '*.bin'), ('All', '*.*')],
            title='저장')
        if path:
            self._write(path)
            self._path = path
            self._path_var.set(path)

    def _write(self, path):
        try:
            with open(path, 'wb') as f:
                for p in self._packets:
                    f.write(p['raw'])
            n_edited = sum(1 for p in self._packets if p['edited'])
            self._status_var.set(f'저장 완료: {path} (편집 {n_edited}개 포함)')
        except OSError as e:
            messagebox.showerror('저장 오류', str(e))

    # ── 내보내기 (CSV / XLSX / PDF) ──────────────────────────────────────

    _EXPORT_HEADERS = ['#', 'tx_abs(ms)', 'tx_time', 'role', 'seq', 'lane',
                       'behavior', 'throttle_pwm', 'steer_pwm', 'hmac', 't_tx(s)']

    def _export_table(self):
        """현재 로드된 패킷 → (헤더, 행리스트). 파싱값을 사람이 읽기 좋은 형태로."""
        rows = []
        for i, p in enumerate(self._packets):
            f = p['fields']
            if f is None:
                rows.append([i + 1, '', '', '', '', '', '', '', '', 'PARSE_ERR', ''])
            else:
                rows.append([
                    i + 1, f['tx_abs'], fmt_ms_of_day(f['tx_abs']),
                    _ROLE_NAMES.get(f['role'], f['role']),
                    f['seq'], f['lane'],
                    _BEH_NAMES.get(f['behavior'], f['behavior']),
                    round(f['throttle_pwm'], 6), round(f['steer_pwm'], 6),
                    'OK' if f['hmac_ok'] else 'BAD',
                    round(f['t_tx'], 6),
                ])
        return self._EXPORT_HEADERS, rows

    def _ask_export_path(self, ext, desc):
        if not self._packets:
            messagebox.showinfo('내보내기', '로드된 패킷이 없습니다.')
            return None
        import os
        init = (os.path.splitext(os.path.basename(self._path))[0] if self._path else 'session')
        return filedialog.asksaveasfilename(
            defaultextension=ext, initialfile=init + ext,
            filetypes=[(desc, '*' + ext), ('All', '*.*')], title=f'{desc} 내보내기')

    def _export_csv(self):
        path = self._ask_export_path('.csv', 'CSV')
        if not path:
            return
        import csv
        headers, rows = self._export_table()
        dlg = _ProgressDialog(self, 'CSV 내보내기', len(rows))
        try:
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:   # BOM → Excel 한글 OK
                w = csv.writer(f)
                w.writerow(headers)
                for i, row in enumerate(rows):
                    w.writerow(row)
                    dlg.advance(i + 1, f'{i + 1} / {len(rows)} 행')
            dlg.done(f'완료 — {len(rows)}행')
        except OSError as e:
            dlg.close(); messagebox.showerror('CSV 오류', str(e)); return
        dlg.close()
        self._status_var.set(f'CSV 내보내기 완료: {path} ({len(rows)}행)')

    def _export_xlsx(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            messagebox.showwarning('XLSX', 'XLSX 내보내기는 openpyxl 이 필요합니다.\n'
                                            '설치: py -m pip install openpyxl')
            return
        path = self._ask_export_path('.xlsx', 'Excel')
        if not path:
            return
        headers, rows = self._export_table()
        dlg = _ProgressDialog(self, 'XLSX 내보내기', len(rows))
        try:
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'packets'
            ws.append(headers)
            hdr_fill = PatternFill('solid', fgColor='DDDDDD')
            for c in ws[1]:
                c.font = Font(bold=True); c.fill = hdr_fill
            ws.freeze_panes = 'A2'
            odd = PatternFill('solid', fgColor='EEF1F5')
            for i, r in enumerate(rows):
                ws.append(r)
                if i % 2 == 1:                  # 교대색 (줄 구분)
                    for c in ws[i + 2]:
                        c.fill = odd
                dlg.advance(i + 1, f'{i + 1} / {len(rows)} 행')
            dlg.set_text('파일 저장 중...')
            wb.save(path)
            dlg.done(f'완료 — {len(rows)}행')
        except OSError as e:
            dlg.close(); messagebox.showerror('XLSX 오류', str(e)); return
        dlg.close()
        self._status_var.set(f'XLSX 내보내기 완료: {path} ({len(rows)}행)')

    def _export_pdf(self):
        try:
            from matplotlib.backends.backend_pdf import PdfPages
            from matplotlib.figure import Figure
        except ImportError:
            messagebox.showwarning('PDF', 'PDF 내보내기는 matplotlib 이 필요합니다.\n'
                                           '설치: py -m pip install matplotlib')
            return
        path = self._ask_export_path('.pdf', 'PDF')
        if not path:
            return
        headers, rows = self._export_table()
        PER_PAGE = 30
        n_pages = max(1, (len(rows) + PER_PAGE - 1) // PER_PAGE)
        dlg = _ProgressDialog(self, 'PDF 내보내기', n_pages)
        try:
            with PdfPages(path) as pdf:
                for pg, start in enumerate(range(0, len(rows), PER_PAGE)):
                    chunk = rows[start:start + PER_PAGE]
                    fig = Figure(figsize=(11.7, 8.3))   # A4 가로
                    ax = fig.add_subplot(111); ax.axis('off')
                    tbl = ax.table(cellText=[[str(c) for c in r] for r in chunk],
                                   colLabels=headers, loc='center', cellLoc='center')
                    tbl.auto_set_font_size(False); tbl.set_fontsize(6.5)
                    tbl.scale(1, 1.3)
                    for (ri, ci), cell in tbl.get_celld().items():
                        if ri == 0:                       # 헤더
                            cell.set_facecolor('#dddddd'); cell.set_text_props(weight='bold')
                        elif ri % 2 == 0:                 # 교대색
                            cell.set_facecolor('#eef1f5')
                    ax.set_title(f'session packets  {start + 1}~{start + len(chunk)} / {len(rows)}',
                                 fontsize=9)
                    pdf.savefig(fig)
                    dlg.advance(pg + 1, f'{pg + 1} / {n_pages} 페이지')
            dlg.done(f'완료 — {n_pages}페이지')
        except OSError as e:
            dlg.close(); messagebox.showerror('PDF 오류', str(e)); return
        dlg.close()
        self._status_var.set(f'PDF 내보내기 완료: {path} ({len(rows)}행)')

    # ── 트리뷰 갱신 ──────────────────────────────────────────────────────

    def _refresh_tree(self):
        self._tree.delete(*self._tree.get_children())
        for i, p in enumerate(self._packets):
            self._tree.insert('', 'end', iid=str(i),
                              values=self._row_values(i, p),
                              tags=self._row_tags(p, i))

    def _row_values(self, i, p):
        f   = p['fields']
        raw = p['raw']
        if f is None:
            return (i + 1, '—\n—', '—\n—', '—\n—', '—\n—', '—\n—',
                    '—\n—', '—\n—', raw.hex(' '))
        return (
            i + 1,
            # 각 셀: hex bytes (위) \n 파싱값 (아래)
            f'{_h(raw, *_SLICES["tx_abs"])}\n{fmt_ms_of_day(f["tx_abs"])}',
            f'{_h(raw, *_SLICES["role"])}\n{_ROLE_NAMES.get(f["role"], str(f["role"]))}',
            f'{_h(raw, *_SLICES["seq"])}\n{f["seq"]}',
            f'{_h(raw, *_SLICES["lane"])}\n{f["lane"]}',
            f'{_h(raw, *_SLICES["behavior"])}\n{_BEH_NAMES.get(f["behavior"], str(f["behavior"]))}',
            f'{_h(raw, *_SLICES["throttle_pwm"])}\n{f["throttle_pwm"]:+.4f}',
            f'{_h(raw, *_SLICES["steer_pwm"])}\n{f["steer_pwm"]:+.4f}',
            _fmt_raw60(raw),
        )

    def _row_tags(self, p, i=0):
        if p['error']:
            return ('error',)
        if p['edited']:
            return ('edited',)
        if not p['fields']['hmac_ok']:
            return ('hmac_bad',)
        return ('evenrow' if i % 2 == 0 else 'oddrow',)   # 정상 행 교대색

    # ── 더블클릭 편집 ────────────────────────────────────────────────────

    def _on_double_click(self, event):
        iid = self._tree.identify_row(event.y)
        if not iid:
            return
        idx = int(iid)
        p   = self._packets[idx]
        if p['error'] or p['fields'] is None:
            messagebox.showwarning('편집 불가',
                                   f'파싱 오류 행은 편집할 수 없습니다.\n{p["error"]}')
            return
        PacketEditDialog(self, idx, dict(p['fields']), self._apply_edit)

    def _apply_edit(self, idx, new_fields):
        try:
            new_raw = _repack(new_fields)
        except Exception as e:
            messagebox.showerror('재패킹 오류', str(e))
            return
        re_fields, err = _parse_one(new_raw)
        p = self._packets[idx]
        p['raw']    = new_raw
        p['fields'] = re_fields
        p['error']  = err
        p['edited'] = True
        self._tree.item(str(idx),
                        values=self._row_values(idx, p),
                        tags=self._row_tags(p, idx))
        n_edited = sum(1 for pkt in self._packets if pkt['edited'])
        self._status_var.set(
            f'{len(self._packets)}개 패킷 | 편집됨 {n_edited}개 (노란색) | 저장 버튼으로 적용')


# ── 편집 팝업 ────────────────────────────────────────────────────────────

class PacketEditDialog(tk.Toplevel):
    """단일 패킷 필드 편집 팝업. 확인 시 on_confirm(idx, new_fields) 호출."""

    def __init__(self, parent, idx, fields, on_confirm):
        super().__init__(parent)
        self.title(f'패킷 편집  —  #{idx + 1}  (seq {fields["seq"]})')
        self.resizable(False, False)
        self._idx       = idx
        self._fields    = fields
        self._on_confirm = on_confirm
        self._build(fields)
        self.grab_set()
        self.focus_set()

    def _build(self, f):
        G = dict(sticky='w', padx=(10, 4), pady=3)

        # 읽기전용 정보 ──────────────────────────────────
        ro = ttk.LabelFrame(self, text='읽기전용 (변경 불가)')
        ro.pack(fill='x', padx=12, pady=(10, 4))
        for i, (label, val) in enumerate([
            ('시각 (tx_abs)',  fmt_ms_of_day(f['tx_abs'])),
            ('seq',            str(f['seq'])),
            ('t_tx (mono)',    f'{f["t_tx"]:.6f} s'),
            ('ver / type',     f'{f["ver"]} / {f["type"]}'),
            ('HMAC',           '✓ 정상' if f['hmac_ok'] else '✗ 불일치 (위변조 의심)'),
        ]):
            tk.Label(ro, text=label, width=18, anchor='w',
                     fg='#666').grid(row=i, column=0, **G)
            tk.Label(ro, text=val,   anchor='w').grid(row=i, column=1, **G)

        # 편집 필드 ──────────────────────────────────────
        ed = ttk.LabelFrame(self, text='편집 가능')
        ed.pack(fill='x', padx=12, pady=4)

        # 각 위젯을 ed 의 자식으로 만들고 grid 배치
        labels = ['역할 (role)', '차선 (lane)', '행동 (behavior)',
                  'throttle_pwm  (-1~1)', 'steer_pwm  (-1~1)']
        for i, lbl in enumerate(labels):
            tk.Label(ed, text=lbl, anchor='w', width=22).grid(
                row=i, column=0, **G)

        self._role_var = tk.StringVar(value=_ROLE_NAMES.get(f['role'], 'LEADER'))
        ttk.Combobox(ed, textvariable=self._role_var,
                     values=[r.name for r in Role],
                     state='readonly', width=16).grid(row=0, column=1, **G)

        self._lane_var = tk.IntVar(value=f['lane'])
        tk.Spinbox(ed, textvariable=self._lane_var,
                   from_=0, to=2, width=8).grid(row=1, column=1, **G)

        self._beh_var = tk.StringVar(value=_BEH_NAMES.get(f['behavior'], 'FOLLOW'))
        ttk.Combobox(ed, textvariable=self._beh_var,
                     values=[b.name for b in DriveBehavior],
                     state='readonly', width=16).grid(row=2, column=1, **G)

        self._thr_var = tk.StringVar(value=f'{f["throttle_pwm"]:.6f}')
        ttk.Entry(ed, textvariable=self._thr_var,
                  width=18).grid(row=3, column=1, **G)

        self._st_var = tk.StringVar(value=f'{f["steer_pwm"]:.6f}')
        ttk.Entry(ed, textvariable=self._st_var,
                  width=18).grid(row=4, column=1, **G)

        # 버튼 ───────────────────────────────────────────
        btn = tk.Frame(self)
        btn.pack(pady=(6, 12))
        ttk.Button(btn, text='확인', command=self._confirm, width=10).pack(side='left', padx=8)
        ttk.Button(btn, text='취소', command=self.destroy,  width=10).pack(side='left', padx=8)

        self.bind('<Return>', lambda _: self._confirm())
        self.bind('<Escape>', lambda _: self.destroy())

    def _confirm(self):
        try:
            thr = float(self._thr_var.get())
            st  = float(self._st_var.get())
        except ValueError:
            messagebox.showerror('입력 오류',
                                 'throttle_pwm / steer_pwm 는 소수 값이어야 합니다.',
                                 parent=self)
            return
        new_fields = dict(self._fields)
        new_fields['role']         = _ROLE_BY_NAME[self._role_var.get()].value
        new_fields['lane']         = int(self._lane_var.get())
        new_fields['behavior']     = _BEH_BY_NAME[self._beh_var.get()].value
        new_fields['throttle_pwm'] = thr
        new_fields['steer_pwm']    = st
        self._on_confirm(self._idx, new_fields)
        self.destroy()


# ── 내보내기 진행 모달 ─────────────────────────────────────────────────────

class _ProgressDialog(tk.Toplevel):
    """내보내기 진행 모달 — 문서 완성까지 프로그레스 바 표시.
    동기 진행(메인 스레드) + 주기적 update_idletasks 로 바를 갱신한다."""

    def __init__(self, parent, title, maximum):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self._max = max(1, int(maximum))
        self._last_pct = -1
        tk.Label(self, text=title, font=('TkDefaultFont', 10, 'bold')).pack(padx=26, pady=(16, 4))
        self._lbl = tk.Label(self, text='준비 중...', fg='#555', width=34)
        self._lbl.pack(padx=26)
        self._bar = ttk.Progressbar(self, mode='determinate', maximum=self._max, length=340)
        self._bar.pack(padx=26, pady=(8, 18))
        self.transient(parent)
        self.protocol('WM_DELETE_WINDOW', lambda: None)   # 완료 전 닫기 방지
        self.grab_set()                                    # 모달
        self._center(parent)
        self.update_idletasks()

    def _center(self, parent):
        try:
            self.update_idletasks()
            px, py = parent.winfo_rootx(), parent.winfo_rooty()
            pw, ph = parent.winfo_width(), parent.winfo_height()
            w, h = self.winfo_reqwidth(), self.winfo_reqheight()
            self.geometry(f'+{px + (pw - w) // 2}+{py + (ph - h) // 2}')
        except tk.TclError:
            pass

    def advance(self, value, text=None):
        """진행값 갱신. ~1% 단위로만 repaint (대량 행에서 성능)."""
        self._bar['value'] = value
        if text is not None:
            self._lbl.config(text=text)
        pct = int(value * 100 / self._max)
        if pct != self._last_pct:
            self._last_pct = pct
            self.update_idletasks()

    def set_text(self, text):
        self._lbl.config(text=text)
        self.update_idletasks()

    def done(self, text='완료'):
        self._bar['value'] = self._max
        self._lbl.config(text=text)
        self.update_idletasks()

    def close(self):
        try:
            self.grab_release()
        except tk.TclError:
            pass
        self.destroy()

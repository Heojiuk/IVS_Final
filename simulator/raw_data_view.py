"""원시 데이터 뷰 — Pi가 dev 모드에서 저장한 .buslog 를 열어 사이클별 전 토픽을 표로 표시 + 내보내기.

  RawDataViewWindow : View 메뉴에서 열리는 Toplevel.
                      .buslog(매 50ms 사이클의 인지·판단·모드·모션·통신 스냅샷)을
                      ttk.Treeview 시간순 표로 보여주고 CSV/XLSX/PDF 로 내보낸다.

파싱은 src/core_module/bus_logger.read_file 재사용(포맷 단일 진실원). 미발행 토픽(valid_mask=0)은
해당 셀을 '—'(표)·빈칸(내보내기)으로 비운다. PDF 는 행 수가 많아 표 대신 리포트형(시계열 plot+요약).
"""
import os
import csv
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import _src_path; _src_path.add()
from core_module.bus_logger import read_file
from core_module.v2v import fmt_ms_of_day
from messages import DriveBehavior

# valid_mask 비트 (bus_logger 와 동일)
_BIT_SCENE, _BIT_CMD, _BIT_MODE, _BIT_EGO, _BIT_PEER, _BIT_LINK = (1 << i for i in range(6))

# 표 컬럼: (col_id, 표시명, 폭, 정렬, 소속섹션, 레코드키)
#   섹션 = valid_mask 비트로 비울지 판단. None=항상 표시. fmt 는 _fmt 에서 처리.
_COLS = [
    ("idx",     "#",        46,  "e",      None,       None),
    ("t_wall",  "시각",      105, "center", None,       "t_abs_ms"),
    # ── 인지 (SCENE) ──
    ("lane",    "차선",      48,  "center", _BIT_SCENE, "current_lane"),
    ("offset",  "offset_cm", 82,  "e",      _BIT_SCENE, "lane_offset_cm"),
    ("heading", "heading",   75,  "e",      _BIT_SCENE, "lane_heading_rad"),
    ("curv",    "curv",      62,  "e",      _BIT_SCENE, "lane_curvature_1pm"),
    ("front",   "front",     58,  "center", _BIT_SCENE, "front_clear"),
    ("dist",    "dist_cm",   68,  "e",      _BIT_SCENE, "dist_front_cm"),
    ("stop",    "stop",      48,  "center", _BIT_SCENE, "stop_signal"),
    # ── 판단 (COMMAND) ──
    ("beh",     "behavior",  102, "center", _BIT_CMD,   "behavior"),
    ("target",  "target",    56,  "center", _BIT_CMD,   "target_lane"),
    # ── 모드 (MODE) ──
    ("mode",    "mode",      82,  "center", _BIT_MODE,  "mode"),
    ("cause",   "cause",     82,  "center", _BIT_MODE,  "cause"),
    # ── 모션 (EGO) ──
    ("thr",     "throttle",  74,  "e",      _BIT_EGO,   "throttle_pwm"),
    ("steer",   "steer",     74,  "e",      _BIT_EGO,   "steer_pwm"),
    ("egobeh",  "ego_beh",   100, "center", _BIT_EGO,   "ego_behavior"),
    # ── 통신 (PEER + LINK) ──
    ("pseq",    "peer_seq",  70,  "e",      _BIT_PEER,  "peer_seq"),
    ("plane",   "p_lane",    52,  "center", _BIT_PEER,  "peer_lane"),
    ("pbeh",    "p_beh",     90,  "center", _BIT_PEER,  "peer_behavior"),
    ("pthr",    "p_thr",     64,  "e",      _BIT_PEER,  "peer_throttle"),
    ("psteer",  "p_steer",   64,  "e",      _BIT_PEER,  "peer_steer"),
    ("link",    "link",      66,  "center", _BIT_LINK,  "link_state"),
    ("age",     "link_age",  72,  "e",      _BIT_LINK,  "link_age_ms"),
    ("lseq",    "last_seq",  62,  "e",      _BIT_LINK,  "link_last_seq"),
]

# 내보내기(CSV/XLSX) 컬럼 = 표와 동일 순서의 (헤더, 레코드키, 섹션비트)
_EXPORT_COLS = [(label, key, sect) for (_cid, label, _w, _a, sect, key) in _COLS if key]


def _fmt(key, val):
    """레코드 값 → 표/내보내기 표시 문자열. dist None 은 빈값으로."""
    if val is None:
        return ""
    if key == "t_abs_ms":
        return fmt_ms_of_day(val)
    if key in ("lane_offset_cm", "lane_heading_rad", "lane_curvature_1pm",
               "dist_front_cm", "throttle_pwm", "steer_pwm",
               "peer_throttle", "peer_steer", "link_age_ms"):
        return f"{val:+.3f}" if key in ("throttle_pwm", "steer_pwm",
                                        "peer_throttle", "peer_steer") else f"{val:.2f}"
    if isinstance(val, bool):
        return "T" if val else "F"
    return str(val)


def _cell(rec, key, sect):
    """레코드 한 칸의 표시값 — 소속 섹션이 valid_mask 에 없으면 미발행이므로 비움."""
    if sect is not None and not (rec["valid_mask"] & sect):
        return None
    return rec.get(key)


# ── 내보내기 (모듈 함수 — GUI 와 분리, 단위 테스트 가능) ─────────────────────

def export_csv(records, path):
    """레코드 → CSV. 미발행 섹션은 빈칸. 첫 컬럼은 사람이 읽는 시각."""
    with open(path, "w", newline="", encoding="utf-8-sig") as f:  # BOM: Excel 한글
        w = csv.writer(f)
        w.writerow([label for label, _k, _s in _EXPORT_COLS])
        for r in records:
            w.writerow([_fmt(k, _cell(r, k, s)) for _label, k, s in _EXPORT_COLS])


def export_xlsx(records, path):
    """레코드 → XLSX (openpyxl). 헤더 굵게·고정, maneuver 행 음영."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "buslog"
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="3949AB")
    lc_fill = PatternFill("solid", fgColor="E3F0FF")   # LANE_CHANGE
    stop_fill = PatternFill("solid", fgColor="FFE0E0")  # STOP

    headers = [label for label, _k, _s in _EXPORT_COLS]
    ws.append(headers)
    for c in ws[1]:
        c.font, c.fill = hdr_font, hdr_fill
        c.alignment = Alignment(horizontal="center")

    for r in records:
        row = []
        for _label, k, s in _EXPORT_COLS:
            v = _cell(r, k, s)
            # 숫자는 숫자 셀로 (Excel 분석 편의), 그 외 문자열
            if v is None:
                row.append(None)
            elif isinstance(v, bool):
                row.append("T" if v else "F")
            elif k == "t_abs_ms":
                row.append(fmt_ms_of_day(v))
            else:
                row.append(v)
        ws.append(row)
        beh = r.get("behavior") if (r["valid_mask"] & _BIT_CMD) else None
        if beh == "LANE_CHANGE":
            for c in ws[ws.max_row]:
                c.fill = lc_fill
        elif beh == "STOP":
            for c in ws[ws.max_row]:
                c.fill = stop_fill

    ws.freeze_panes = "A2"
    for i, (_label, _k, _s) in enumerate(_EXPORT_COLS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 12
    wb.save(path)


def export_pdf(records, path, src_name=""):
    """레코드 → 리포트형 PDF (matplotlib). 시계열 plot 4종 + 요약 텍스트.
    행이 수천 개일 수 있어 전체 표 대신 리포트로 요약."""
    from matplotlib.figure import Figure
    import matplotlib
    matplotlib.rcParams["font.sans-serif"] = ["Malgun Gothic", "DejaVu Sans"]
    matplotlib.rcParams["axes.unicode_minus"] = False

    if not records:
        raise ValueError("레코드 없음")
    t0 = records[0]["t_abs_ms"]
    ts = [(r["t_abs_ms"] - t0) / 1000.0 for r in records]   # 시작 기준 초

    def col(key, sect):
        nan = float("nan")
        out = []
        for r in records:
            if sect is not None and not (r["valid_mask"] & sect):
                out.append(nan); continue
            v = r.get(key)
            out.append(nan if v is None else v)
        return out

    thr = col("throttle_pwm", _BIT_EGO)
    steer = col("steer_pwm", _BIT_EGO)
    lane = col("current_lane", _BIT_SCENE)
    beh_num = []
    for r in records:
        if r["valid_mask"] & _BIT_CMD:
            nm = r.get("behavior")
            beh_num.append(DriveBehavior[nm].value if nm in DriveBehavior.__members__
                           else float("nan"))
        else:
            beh_num.append(float("nan"))

    fig = Figure(figsize=(11, 7.2), dpi=110)
    specs = [("throttle_pwm", thr, "#1155cc"), ("steer_pwm", steer, "#cc5511"),
             ("current_lane (인지)", lane, "#118811"), ("behavior (판단)", beh_num, "#8811aa")]
    for i, (title, ys, color) in enumerate(specs):
        ax = fig.add_subplot(3, 2, i + 1)
        ax.plot(ts, ys, linewidth=1, color=color, drawstyle="steps-post"
                if "lane" in title or "behavior" in title else "default")
        ax.set_title(title, fontsize=10)
        ax.set_xlabel("t (s)", fontsize=8)
        ax.grid(True, alpha=0.3)
        ax.tick_params(labelsize=7)
        if "behavior" in title:
            ax.set_yticks([b.value for b in DriveBehavior])
            ax.set_yticklabels([b.name for b in DriveBehavior], fontsize=6)

    # 요약 텍스트 (하단 2칸 병합 영역)
    ax = fig.add_subplot(3, 1, 3)
    ax.axis("off")
    lines = [f"파일: {src_name}",
             f"사이클: {len(records)}    구간: {fmt_ms_of_day(t0)} ~ "
             f"{fmt_ms_of_day(records[-1]['t_abs_ms'])}  ({ts[-1]:.1f}s)"]
    # behavior 전이
    prev, trans = None, []
    for r in records:
        b = r.get("behavior") if (r["valid_mask"] & _BIT_CMD) else None
        if b != prev and b is not None:
            trans.append(f"{fmt_ms_of_day(r['t_abs_ms'])} {prev}->{b}")
            prev = b
    lines.append("behavior 전이: " + ("  ".join(trans) if trans else "없음"))
    from collections import Counter
    link = Counter(r["link_state"] for r in records if (r["valid_mask"] & _BIT_LINK))
    lines.append(f"링크 상태: {dict(link) if link else '데이터 없음'}")
    ax.text(0.01, 0.95, "\n".join(lines), va="top", ha="left", fontsize=9,
            transform=ax.transAxes, wrap=True)   # 기본 sans-serif(Malgun Gothic) — 한글 렌더

    fig.suptitle("Raw Data View 리포트", fontsize=13, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 1, 0.97))
    fig.savefig(path)


# ── GUI 창 ────────────────────────────────────────────────────────────────────
class RawDataViewWindow(tk.Toplevel):
    """저장된 .buslog 를 열어 사이클별 전 토픽을 표로 보고 CSV/XLSX/PDF 로 내보낸다."""

    def __init__(self, parent, initial_dir):
        super().__init__(parent)
        self.title("원시 데이터 뷰 — Raw Data View")
        self.geometry("1300x620")
        # DevBus 폴더가 있으면 거기서 시작
        devbus = os.path.join(initial_dir, "DevBus")
        self._initial_dir = devbus if os.path.isdir(devbus) else initial_dir
        self._records = []
        self._path = None

        bar = ttk.Frame(self); bar.pack(fill="x", padx=4, pady=4)
        ttk.Button(bar, text="📂 .buslog 열기", command=self._open).pack(side="left")
        ttk.Separator(bar, orient="vertical").pack(side="left", fill="y", padx=6)
        ttk.Label(bar, text="내보내기:").pack(side="left")
        ttk.Button(bar, text="CSV",  command=lambda: self._export("csv")).pack(side="left", padx=1)
        ttk.Button(bar, text="XLSX", command=lambda: self._export("xlsx")).pack(side="left", padx=1)
        ttk.Button(bar, text="PDF",  command=lambda: self._export("pdf")).pack(side="left", padx=1)
        self._info = tk.StringVar(value=".buslog 파일을 여세요.")
        ttk.Label(bar, textvariable=self._info, foreground="#1155cc").pack(side="left", padx=8)

        frm = ttk.Frame(self); frm.pack(fill="both", expand=True, padx=4, pady=(0, 4))
        col_ids = [c[0] for c in _COLS]
        _s = ttk.Style()
        _s.configure("RDV.Treeview", rowheight=22, font=("Consolas", 9))
        _s.configure("RDV.Treeview.Heading", font=("TkDefaultFont", 8, "bold"))
        self._tree = ttk.Treeview(frm, columns=col_ids, show="headings",
                                  selectmode="browse", style="RDV.Treeview")
        for cid, label, width, anchor, _sect, _key in _COLS:
            self._tree.heading(cid, text=label)
            self._tree.column(cid, width=width, minwidth=30, anchor=anchor, stretch=False)
        vsb = ttk.Scrollbar(frm, orient="vertical", command=self._tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        frm.rowconfigure(0, weight=1)
        frm.columnconfigure(0, weight=1)
        self._tree.tag_configure("evenrow", background="#ffffff")
        self._tree.tag_configure("oddrow", background="#eef1f5")
        self._tree.tag_configure("lc", background="#e3f0ff")     # LANE_CHANGE
        self._tree.tag_configure("stop", background="#ffe0e0")   # STOP

    def _open(self):
        path = filedialog.askopenfilename(
            initialdir=self._initial_dir,
            filetypes=[("Bus log", "*.buslog"), ("All", "*.*")],
            title=".buslog 열기", parent=self)
        if not path:
            return
        try:
            header, records = read_file(path)
        except Exception as e:
            messagebox.showerror("원시 데이터 뷰", f"파싱 실패: {e}", parent=self)
            return
        self._path, self._records = path, records
        n = len(records)
        dur = (records[-1]["t_abs_ms"] - records[0]["t_abs_ms"]) / 1000.0 if n else 0.0
        self._info.set(f"{os.path.basename(path)}  |  role={header['role_name']}  "
                       f"{n}사이클  {dur:.1f}s")
        self._fill_table()

    def _fill_table(self):
        self._tree.delete(*self._tree.get_children())
        for i, r in enumerate(self._records):
            vals = [str(i)]
            for _cid, _label, _w, _a, sect, key in _COLS[1:]:
                vals.append(_fmt(key, _cell(r, key, sect)) or "—")
            beh = r.get("behavior") if (r["valid_mask"] & _BIT_CMD) else None
            tag = "lc" if beh == "LANE_CHANGE" else "stop" if beh == "STOP" else \
                  ("evenrow" if i % 2 == 0 else "oddrow")
            self._tree.insert("", "end", iid=str(i), values=vals, tags=(tag,))

    def _export(self, fmt):
        if not self._records:
            messagebox.showinfo("내보내기", "먼저 .buslog 를 여세요.", parent=self)
            return
        base = os.path.splitext(self._path)[0]
        ext = {"csv": ".csv", "xlsx": ".xlsx", "pdf": ".pdf"}[fmt]
        types = {"csv": [("CSV", "*.csv")], "xlsx": [("Excel", "*.xlsx")],
                 "pdf": [("PDF", "*.pdf")]}[fmt]
        out = filedialog.asksaveasfilename(
            initialfile=os.path.basename(base) + ext, defaultextension=ext,
            initialdir=os.path.dirname(self._path), filetypes=types,
            title=f"{fmt.upper()} 내보내기", parent=self)
        if not out:
            return
        try:
            if fmt == "csv":
                export_csv(self._records, out)
            elif fmt == "xlsx":
                export_xlsx(self._records, out)
            else:
                export_pdf(self._records, out, os.path.basename(self._path))
        except Exception as e:
            messagebox.showerror("내보내기", f"{fmt.upper()} 실패: {e}", parent=self)
            return
        messagebox.showinfo("내보내기", f"저장 완료:\n{out}", parent=self)
